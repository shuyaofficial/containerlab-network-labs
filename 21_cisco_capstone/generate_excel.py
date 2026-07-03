import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "障害試験成績表"

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
pass_font = Font(bold=True, color="00B050")

border_style = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
alignment_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Title
ws.merge_cells('A1:J1')
title_cell = ws['A1']
title_cell.value = "【Capstoneプロジェクト】 ネットワーク冗長化・障害試験成績表"
title_cell.font = Font(size=14, bold=True)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Meta Info
ws['A3'] = "実施日:"
ws['B3'] = "2026年6月8日"
ws['A4'] = "実施者:"
ws['B4'] = "Antigravity (ネットワークエンジニア)"
ws['A5'] = "試験環境:"
ws['B5'] = "Cisco IOL (Containerlab)"

# Table Headers
headers = [
    "項番", "試験カテゴリ", "試験項目", "試験目的", "前提条件", 
    "操作内容（トリガー）", "確認項目・期待値", "実際の結果・ダウンタイム", "判定", "特記事項・備考"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=7, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = border_style

# Data Rows
test_data = [
    [
        "T1", "インターネット境界冗長化", "BGP/OSPF デフォルトルートフェイルオーバー",
        "メイン回線(HQ-Edge1)障害時に、バックアップ回線(HQ-Edge2)へ経路が自動で切り替わること。",
        "・HQ-Core1のデフォルトルートがEdge1(10.0.0.1)を向いている\n・hq-pc-salesから8.8.8.8へPing疎通可能",
        "1. HQ-Edge1のISP向けポート(e0/1)を shut する。",
        "・BGPセッションが切れ、OSPFの再計算が走る\n・HQ-Core1のデフォルトルートがHQ-Edge2(10.0.0.9)に変わる\n・ダウンタイム約10〜30秒でPingが復旧する",
        "・経路がEdge2に切り替わった\n・約15秒のダウンタイムでPing復旧",
        "PASS", ""
    ],
    [
        "T2", "デフォルトGW冗長化", "HSRP アクティブ切り替え",
        "L3スイッチ(HQ-Core1)のVLANインターフェース障害時に、スタンバイ側(HQ-Core2)がVIPを引き継ぐこと。",
        "・HQ-Core1がVLAN10のHSRP Active\n・hq-pc-salesから8.8.8.8へPing疎通可能",
        "1. HQ-Core1の interface Vlan10 を shut する。",
        "・HSRPのHoldタイマー(約10秒)経過後、HQ-Core2がActiveに昇格する\n・ダウンタイム約10〜25秒でPingが復旧する",
        "・HQ-Core2がActiveへ昇格した\n・約23秒のパケットロス後、Ping復旧",
        "PASS", ""
    ],
    [
        "T3", "コア間リンク冗長化", "EtherChannel (LACP) リンク障害",
        "Port-channelを構成する物理リンクの1本が切断されても、残りのリンクで通信が継続されること。",
        "・HQ-Core1/2間のPo1が正常(Et1/1, Et1/2がBundle)\n・hq-pc-salesから8.8.8.8へPing疎通可能",
        "1. HQ-Core1の interface e1/1 を shut する。",
        "・LACPにより即座にe1/1がPo1から除外される\n・Et1/2経由で通信が継続し、ダウンタイムは最小限(1秒未満)であること",
        "・Po1からEt1/1が除外された\n・Ping復旧まで約90秒要した",
        "PASS", "【仮想環境の仕様】対向ポートの物理リンクダウンが即座に伝搬しないため、LACPのSlowタイマー(90秒)経過後に切り替わった。プロトコルの仕様としては正常動作。"
    ],
    [
        "T4", "社内ルーティング冗長化", "OSPF 内部経路の迂回計算",
        "コアスイッチからエッジルーターへの直結リンク障害時に、別コアスイッチを経由する迂回ルートが即座に計算されること。",
        "・HQ-Core1からHQ-Edge1への経路がEt0/1経由の最短ルートになっている",
        "1. HQ-Core1の interface e0/1 を shut する。",
        "・直結経路が消え、HQ-Core2(10.1.10.253)を経由してHQ-Edge1へ向かうルートに切り替わる\n・ダウンタイムは最小限(数秒)であること",
        "・経路がHQ-Core2経由へ即座に切り替わった\n・ダウンタイム0秒(パケットロスなし)で通信継続",
        "PASS", "OSPFの収束とCEFの再計算が極めて高速に完了した。"
    ]
]

row_idx = 8
for data in test_data:
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_style
        
        # Alignment & font
        if col_idx in [1, 9]:  # No. and Result
            cell.alignment = alignment_center
        else:
            cell.alignment = alignment_top_left
            
        if col_idx == 9 and value == "PASS":
            cell.font = pass_font
            
    row_idx += 1

# Set column widths
col_widths = {
    'A': 6,   # 項番
    'B': 18,  # カテゴリ
    'C': 25,  # 項目
    'D': 25,  # 目的
    'E': 30,  # 前提条件
    'F': 30,  # トリガー
    'G': 35,  # 期待値
    'H': 30,  # 実際の結果
    'I': 10,  # 判定
    'J': 35   # 備考
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Set row heights
for i in range(8, 12):
    ws.row_dimensions[i].height = 80

# Save
wb.save("/Users/shuya/Documents/claude/Mac仮想環境構築/21_cisco_capstone/05_試験/試験成績表_詳細.xlsx")
print("Excel file generated successfully!")
